import pandas as pd # I really wanted to call this catBear but will stay with normal naming conventions...iykyk ;)
import openpyxl
import time
import datetime
import numpy as np
import os
from math import floor
import shutil
sourcelocation = "Templates/P3MasterTemplate6.2.23.xlsx"
destinationlocation = "Part3V1.xlsx"
shutil.copy(sourcelocation, destinationlocation)
processedData = pd.ExcelFile("../../../Documents/Github/Automated-Stock-Selector/ProcessedSheets/June2023/ProcessedDataV2.xlsx")
writeTo = pd.ExcelWriter(destinationlocation)
core1 = openpyxl.load_workbook(sourcelocation)


sheetNames = processedData.sheet_names
stoppingIndecies = [702, 1220, 1053, 73, 45, 148, 175, 792, 261, 182, 1498, 68, 680]
elimIndex = 2

def main():
    global currSheet, elimIndex
    for currSheet in sheetNames:
        sheetIndex = sheetNames.index(currSheet)
        currCounter = 0
        rowCounter = 3
        tempWorksheet = core1[currSheet]  
        if currSheet == 'Miscellaneous':
            selectedSheet = pd.read_excel(processedData, currSheet, header = 1)

            rev2Sheet = pd.read_excel(sourcelocation, currSheet, header = 1)
            for row in selectedSheet.iterrows(): # Testing
                continueRunning = True
                currRow = selectedSheet.iloc[currCounter].fillna(-100000).to_numpy()
                continueRunning = revProcessor(currRow)
                if continueRunning:
                    ltoA = round(float(int(selectedSheet['Tot. Liabilities'][currCounter].replace(',', ''))/\
                        int(selectedSheet['Tot. Asset'][currCounter].replace(',', ''))), 2)
                    numemployee = selectedSheet['# FT Workers'][currCounter].replace(',', '')
                    corpName = selectedSheet['Company Name'][currCounter]
                    corpTick = selectedSheet['Tick Symbol'][currCounter]
                    corpCountry = selectedSheet['Country'][currCounter]
                    ipoYR = selectedSheet['IPO Year'][currCounter]
                    corpEX = selectedSheet['Exchange'][currCounter]
                    entPVal = selectedSheet['Enterprise Val.'][currCounter]
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('L/A Ratio') + 1).value = ltoA
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('# FT Workers') + 1).value = numemployee
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('Enterprise Val.') + 1).value = entPVal
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('Company Name') + 1).value = corpName
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('Tick Symbol') + 1).value = corpTick
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('Country') + 1).value = corpCountry
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('IPO Year') + 1).value = ipoYR
                    tempWorksheet.cell(row = rowCounter, column = rev2Sheet.columns.get_loc('Exchange') + 1).value = corpEX
                # print("XX")
                    rowCounter += 1

                currCounter += 1
    core1.save(filename=destinationlocation)

    # writeTo._save
    # writeTo.close()  

def revProcessor(row):
    revErr = 100
    arr = replw0(row[286:304], 0)
    if np.count_nonzero(arr) < 3: return errorHandler(revErr, row)
    elif arr[2] == 0: return errorHandler(revErr + 1, row)
    elif arr[0] == arr[1] == arr[2] == 0: return errorHandler(revErr + 2, row)
    elif arr[0] == 0 and arr[1] == 0: 
        if np.average(arr[2:7]) < 50: return errorHandler(revErr + 3, row)
    else:
        if np.average(np.max(arr[0:2]) + arr[2:6]) < 50: return errorHandler(revErr + 3, row)

    lastInd = 2
    if arr[0] != 0: arr[0] = round(percentChange(arr[0], arr[2]),3)
    if arr[1] != 0: arr[1] = round(percentChange(arr[1], arr[2]),3)
    for i in range(3, len(arr)):
        if arr[i] != 0:
            arr[lastInd] = round(percentChange(arr[lastInd], arr[i]),2)
            lastInd = i
    return arr, True

def percentChange(final, initial):
    return (final - initial)/initial * 100

def replw0(arr, element):
    for i in range (0, len(arr)):
        if arr[i] == -100000:
            arr[i] = element
    return arr

def calculateReturnScore():

    pass

def errorHandler(flag, currRow):
    global elimIndex
    currRow = replw0(currRow, None)
    tempWKST = core1['ELIMINATED']
    flagComment = getErrorCode(flag)
    for i in range(1,7):
        x = i
        if i == 5: 
            tempWKST.cell(row = elimIndex, column = i).value = currSheet
        else:
            if i > 5: x -= 1
            tempWKST.cell(row = elimIndex, column = i).value = currRow[x - 1]
    tempWKST.cell(row = elimIndex, column = 7).value = flagComment
    core1.save(destinationlocation)
    elimIndex += 1
    return False;

def getErrorCode(input):
    switch = {
        100: "E100: Less than 3 valid revenue entries",
        101: "E101: 2022 Revenue Missing",
        102: "E102: First 3 revenue values are all missing or 0",
        103: "E103: Revenue average is below $50 Million ",
        }
    return switch.get(input, "")

if __name__ == '__main__':
    main()