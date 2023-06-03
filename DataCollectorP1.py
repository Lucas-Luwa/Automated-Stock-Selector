# Import libraries
import requests
import time
from bs4 import BeautifulSoup
import openpyxl
import time
import datetime
import calendar

#Modify Toggles 
nasdaqActive, nyseActive, recoveryMode, nasdaqTestMode, nyseTestMode, manualResetVerNum, manualStartIndex, generateSheet  = True, True, False, False, False, False, False, True
recoveryIndecies = [424, 1101, 564, 40, 29, 88, 59, 596, 87, 52, 799, 17, 212]
recoveryFileName = "May23RawData CompleteBackup Nasdaq.xlsx"
recoveryFailureIndex = 631 #Use this for manual index start as well 
nasdaqRecoveryIndex = 2279 #Use this for manual index start as well 
testModeStocks = 3
nyseRecoveryIndex = 2 #Use this for manual index start as well
customVersionNumber = 8

def main():
    global start, sheetNames, core1, newFileName
    start = time.time() 
    numNasdaqStocks, numNYSEStocks, versionNumber, year, month = startupSequence()
    print("Starting program...")
    nasdaqName = "CoreExcelFiles/NASDAQ2.4.23.xlsx"
    nyseData = "CoreExcelFiles/NYSE5.26.23.xlsx"
    coreName = "CoreExcelFiles/P1MasterTemplate6.2.23.xlsx"  if not recoveryMode and nasdaqActive else recoveryFileName
    nasdaq1 = openpyxl.load_workbook(nasdaqName)
    nyse1 = openpyxl.load_workbook(nyseData)
    core1 = openpyxl.load_workbook(coreName)
    nasdaq = nasdaq1.active
    nyse = nyse1.active
    sheetNames = core1.sheetnames
    rowIndecies = [3] * (len(sheetNames) - 1) if not recoveryMode and nasdaqActive else recoveryIndecies
    genXlSheets(nasdaqName, coreName) #It's the same for both.

    failedIndex = 2 if not recoveryMode and nasdaqActive else recoveryFailureIndex
    startIndexNasdaq = 2 if not recoveryMode and not manualStartIndex else nasdaqRecoveryIndex
    startIndexNYSE = 2 if not recoveryMode and not manualStartIndex else nyseRecoveryIndex
    
    newFileName = str(calendar.month_name[int(month)]) + str(year) + "RawDataV" + str(versionNumber) + ".xlsx"
    if nasdaqActive: rowIndecies, failedIndex = excelWriter(nasdaq, rowIndecies, failedIndex, numNasdaqStocks, 1, startIndexNasdaq)
    if nyseActive: rowIndecies, failedIndex = excelWriter(nyse, rowIndecies, failedIndex, numNYSEStocks, 2, startIndexNYSE)
    versionUpdate(month, year, versionNumber)

    end = time.time()
    elapsed = round(end - start)
    print("Total Time Elapsed: ", str(datetime.timedelta(seconds = elapsed)))

def startupSequence():
    numNasdaqStocks = testModeStocks if nasdaqTestMode else 4659
    numNYSEStocks = testModeStocks if nyseTestMode else 2949
    year, month, day = str(datetime.date.today()).split('-')
    recoveryFile = open('Recovery.txt')
    recoveryLines = recoveryFile.readlines()
    versionNumber = 0
    if len(recoveryLines) == 7:
        existingMonthYear = recoveryLines[4].split(" ")
        existingVersionNum = recoveryLines[5].split(" ")
        if not manualResetVerNum: versionNumber = int(existingVersionNum[1]) + 1 if (str(calendar.month_name[int(month)] + str(year))) == str(existingMonthYear[1]).strip() else 0
        else: versionNumber = customVersionNumber
    return numNasdaqStocks, numNYSEStocks, versionNumber, year, month

def versionUpdate(month, year, versionNumber):
    recoveryFile = open("Recovery.txt","a")
    recoveryFile.write("\nMonth/Year: "+ str(calendar.month_name[int(month)]) + str(year))
    recoveryFile.write("\nVersionNumber: "+ str(versionNumber))
    recoveryFile.write("\nUpdated: "+ str(datetime.date.today()))
    recoveryFile.close()


def excelWriter(sheet, rowIndecies, failedIndex, endVal, selectorBit, startIndex):
    counter = startIndex - 2;
    if manualStartIndex: endVal = startIndex + endVal
    for row in sheet.iter_rows(startIndex, endVal):
        start1 = time.time()
        ticker, name, country, ipoyr, currSheet, industry = row[0].value, row[1].value, row[6].value, row[7].value, row[9].value, row[10].value
        processed = dataCollect(ticker)
        if processed.__contains__('500: Internal Server Error') and len(processed) < 100:
            tempWKST = core1['FAILS']
            if selectorBit == 1: initialize = [ticker, name, country, ipoyr, industry, "NASDAQ"]
            if selectorBit == 2: initialize = [ticker, name, country, ipoyr, industry, "NYSE"]
            for i in range (1, 7):
                tempWKST.cell(row = failedIndex, column = i).value = initialize[i - 1]
            failedIndex += 1
        else:
            Series1, Series2, Series3 = dataSplitter(processed)
            if (currSheet == None): currSheet = 'Unspecified'
            sheetIndex = sheetNames.index(currSheet)
            tempWorksheet = core1[currSheet]  
            #Tags
            if selectorBit == 1: initialize = [name, ticker, country, ipoyr, "NASDAQ"]
            if selectorBit == 2: initialize = [name, ticker, country, ipoyr, "NYSE"]
            for i in range (1, (6 + len(Series1) + len(Series2) + len(Series3))):
                if i <= 5: tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = initialize[i - 1]
                if i > 5 and i <= (5 + len(Series1)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series1[i - 6]
                if i > (5 + len(Series1)) and i <= (5 + len(Series1) + len(Series2)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series2[i - 6 - len(Series1)]
                if i > (5 + len(Series1) + len(Series2)) and i <= (5 + len(Series1) + len(Series2) + len(Series3)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series3[i - 6 - len(Series1) - len(Series2)]

            rowIndecies[sheetIndex] += 1
        if generateSheet: core1.save(newFileName)
        counter += 1
        # print(Series2)
        #System Error Recovery
        recoveryFile = open("Recovery.txt","w")
        recoveryFile.write("rowIndecies: " + str(rowIndecies) + "\n")
        recoveryFile.write("failedIndex: " + str(failedIndex) + "\n")
        recoveryFile.write("Last successful Ticker: " + str(ticker) + "\n")
        recoveryFile.write("Current Exchange: NASDAQ") if selectorBit == 1 else recoveryFile.write("Current Exchange: NYSE")
        recoveryFile.close()
        end = time.time()

        #Impatient programmer pacifier 
        #Change the counter % [Some number] to change number of statements printed
        if (counter % 1 == 0 or counter == 1 or counter == endVal - 1): 
            print("Running ", counter, " of ", endVal - 1, " in NASDAQ ", "| ", ticker, " | Time Elapsed: ", str(datetime.timedelta(seconds = round(end - start1))), \
            " | Cumulative Time Elapsed", str(datetime.timedelta(seconds = round(end - start)))) \
            if selectorBit == 1 else print("Running ", counter, " of ", endVal - 1, " in NYSE ",  "| ", ticker, "| Time Elapsed: ", str(datetime.timedelta(seconds = round(end - start1))), \
            " | Cumulative Time Elapsed", str(datetime.timedelta(seconds = round(end - start))))
    print("NASDAQ Complete \n") if selectorBit == 1 else print("NYSE Complete \n")
    return rowIndecies, failedIndex

def dataSplitter(input):
    delimiters = ['P/E', 'Forward P/E', 'P/E to S&P500', 'Market CAP', 'Div Yield', '% Held by Insiders', '% Held by Institutions', 
    'Beta', 'PEG', '52w. high/low', 'Avg. Daily Volume', 'Return']
    charOffset = [0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 1]
    inputoffset = [0, 0, 13, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    dotFlag = [False, False, False, False, False, False, False, False, False, False, True, False]
    initialSetRet = list()
    for i in range(len(delimiters)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters[i])], charOffset[i] , dotFlag[i])
        initialSetRet.append(currVal)
        input = input[(indexLoc(input, delimiters[i])+inputoffset[i]):]
    input = input[indexLoc(input, 'linearHighLow'):] # Discard useless data prior to this point

    #ISR contains PE, Forward PE, PE to SP500, MarketCAP(B/T), Dividend, Shares Shorted, Percent Insiders, 
    #Percent Institutions, Beta, Peg Ratio, 52wk h/l, Avg daily shares traded {Additional Definitions at Bottom}
    # numericalDataProcessor(input)

    secondaryRetSet = list()
    highlow = findBackNumber(input[:indexLoc(input, 'Currency:')], 0 , True)
    secondaryRetSet.append(highlow)
    # print(input)

    yearsTTM = findBackNumber(input[:indexLoc(input, 'Revenue')], 3 , True)
    secondaryRetSet.append(yearsTTM)
    delimiters2 = ['Earnings per share', 'FCF per share', 'Dividends per share', 'CAPEX per share', 'Book Value per sh.', 'Comm.Shares outs.'
    , 'P/E to S&P500', 'Avg. annual div. yield', 'Revenue (m)', 'Operating margin', 'Depreciation (m)', 'Net profit (m)', 'Net profit margin', 'Working capital (m)', 'ROIC'
    , 'Return on capital', 'Return on equity', 'Plowback ratio', 'Div.&Repurch./FCF']
    inputoffset2 = [0, 0, 0, 0, 0, 0, 13, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(len(delimiters2)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters2[i])], 0 , True)
        secondaryRetSet.append(currVal)
        input = input[(indexLoc(input, delimiters2[i])+inputoffset2[i]):]
    input = input[indexLoc(input, 'Total liabilities'):]# Discard useless data prior to this point

    #highlow, yrs, Revenue Per Share, Earnings Per Share, FCF Per Share, Dividends, Capex, Book Value/Share, PE Ratio, PE-500, DivYield
    #Revenue, Operating Margin, Net Profit, Net Profit Margin, Equity, ROIC

    finalset = list()
    delimiters3 = ['Total assets', 'Long-term debt', 'Cash and equiv.', 'Goodwill', 'Common stock', 'Working Capital']
    for i in range(len(delimiters3)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters3[i])], 0 , False)
        finalset.append(currVal)
        input = input[indexLoc(input, delimiters3[i]):]
    input = input[indexLoc(input, 'Full-time employees: '):]
    numMinions = findBackNumber(input[:], 0, True)
    finalset.append(numMinions)

    #Total Liabilities, Total Assets, Long-Term Debt, Cash & Equiv., Retained Earnings, Enterprise Value, number of FT workers

    return [initialSetRet, secondaryRetSet, finalset]


def dataCollect(input):
    url = "https://roic.ai/company/" + str(input)
    response = requests.get(url)
    contents = BeautifulSoup(response.text, 'html.parser')
    splitter = 'City'
    clipped = contents.get_text().split(splitter)
    return clipped[0]

def genXlSheets(input1, writeloc):
    source = openpyxl.load_workbook(input1)
    e1r = source.active
    writeTo = openpyxl.load_workbook(writeloc)
    target = writeTo.active
    setholder = set()
    #Finding Names of Sections
    for row in e1r.iter_rows(2, e1r.max_row + 1):
        setholder.add(row[9].value)

    for element in setholder:
        if element in writeTo.sheetnames:
            pass
        elif element == None and 'Unspecified' not in writeTo.sheetnames:
            writeTo.create_sheet('Unspecified')
            writeTo.save(writeloc)
        elif element != None:
            writeTo.create_sheet(element)
            writeTo.save(writeloc)

def indexLoc(input, target):
    location = input.find(target)
    return location

def findBackNumber(input, offset, disableDots):
    returnVal = str()
    dotcounter = 0
    input = input.replace("(Infinity)", "{0}")
    input = input.replace("Infinity", "{0}")
    for i in range(len(input) - 1, -1, -1):
        approvedCharacters = {'-', ',', '%', ' ', '/', '(', ')', '{', '}'}
        if(input[i].isnumeric() or offset > 0 or input[i] in approvedCharacters):
            returnVal = input[i] + returnVal
            if (input[i] != '-' and input[i] != ',' and input[i]!= '%' and input[i]!= ' '):
                offset -= 1
        elif (input[i] == '.' and dotcounter == 0):
            returnVal = input[i] + returnVal
            if (disableDots != True):
                dotcounter += 1
        else:
            break
    return returnVal 

if __name__ == "__main__":
    main()


#PEG Ratio: P/E Ratio OVER Earnings Growth Rate
#Beta Number: How stock behaves relative to others. >1 is risky <1 is safe. Unused
