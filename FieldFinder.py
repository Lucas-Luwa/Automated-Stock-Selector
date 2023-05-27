import openpyxl

nasdaq = openpyxl.load_workbook("NASDAQ Ticker 2.4.23.xlsx")
nyse = openpyxl.load_workbook("NASDAQ Ticker 2.4.23.xlsx")
nasdaqTick = nasdaq.active
nyseTick = nyse.active

nasdaqHolder = set()
nyseHolder = set()
for row in nasdaqTick.iter_rows(2, nasdaqTick.max_row + 1):
    nasdaqHolder.add(row[9].value)

for row in nyseTick.iter_rows(2, nyseTick.max_row + 1):
    nyseHolder.add(row[9].value)

for element in nasdaqHolder:
    print(element)
    
print("NYSE")
for element in nyseHolder:
    print(element)