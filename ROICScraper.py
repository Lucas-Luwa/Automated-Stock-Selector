# Import libraries
import requests
import time
from bs4 import BeautifulSoup
import openpyxl
import time
import ExcelProcessor
import datetime

nasdaqActive, nyseActive, recoveryMode = False, True, False

def main():
    start = time.time()
    print("Starting program...")
    nasdaqName = "NASDAQ Ticker 2.4.23.xlsx"
    nyseData = "NYSE Ticker 5.26.23.xlsx"
    coreName = "MasterTemplate Updated May 2023.xlsx"  if not recoveryMode else coreName = "May23RawData Recovery1 5.27.23.xlsx"
    nasdaq1 = openpyxl.load_workbook(nasdaqName)
    nyse1 = openpyxl.load_workbook(nyseData)
    core1 = openpyxl.load_workbook(coreName)
    nasdaq = nasdaq1.active
    nyse = nyse1.active
    sheetNames = core1.sheetnames
    #Remember to replace recovery values with your own values
    rowIndecies = [3] * (len(sheetNames) - 1) if not recoveryMode else rowIndecies = [383, 1014, 512, 33, 27, 74, 53, 535, 77, 47, 726, 15, 190]
    genXlSheets(nasdaqName, coreName) #It's the same for both.

    #Change the following values for recovery mode.
    failedIndex = 2 if not recoveryMode else failedIndex = 591
    startIndexNasdaq = 2 if not recoveryMode else startIndexNasdaq = 4238
    startIndexNYSE = 2 if not recoveryMode else startIndexNYSE = 27

    if nasdaqActive: rowIndecies, failedIndex = excelWriter(core1, nasdaq, rowIndecies, sheetNames, failedIndex, 4659, 1, startIndexNasdaq, start) #4659
    if nyseActive: rowIndecies, failedIndex = excelWriter(core1, nyse, rowIndecies, sheetNames, failedIndex, 2949, 2, startIndexNYSE, start) #2949
    end = time.time()
    elapsed = round(end - start)
    print("\nTotal Time elapsed: ", str(datetime.timedelta(seconds = elapsed)))

def excelWriter(core1, sheet, rowIndecies, sheetNames, failedIndex, endVal, selectorBit, startIndex, processStartTime):
    counter = startIndex - 2;
    for row in sheet.iter_rows(startIndex, endVal): #Replace with 4659 for all
        start = time.time()
        ticker, name, country, ipoyr, currSheet, industry = row[0].value, row[1].value, row[6].value, row[7].value, row[9].value, row[10].value
        processed = dataCollect(ticker)
        if processed.__contains__('500: Internal Server Error') and len(processed) < 100:
            tempWKST = core1['FAILS']
            if selectorBit == 1: initialize = [ticker, name, country, ipoyr, industry, "NASDAQ"]
            if selectorBit == 2: initialize = [ticker, name, country, ipoyr, industry, "NYSE"]
            for i in range (1, 7):
                tempWKST.cell(row = failedIndex, column = i).value = initialize[i - 1]
            failedIndex += 1
        else:
            Series1, Series2, Series3 = dataSplitter(processed)
            if (currSheet == None): currSheet = 'Unspecified'
            sheetIndex = sheetNames.index(currSheet)
            tempWorksheet = core1[currSheet]  
            #Tags
            if selectorBit == 1: initialize = [name, ticker, country, ipoyr, "NASDAQ"]
            if selectorBit == 2: initialize = [name, ticker, country, ipoyr, "NYSE"]
            for i in range (1, (6 + len(Series1) + len(Series2) + len(Series3))):
                if i <= 5: tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = initialize[i - 1]
                if i > 5 and i <= (5 + len(Series1)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series1[i - 6]
                if i > (5 + len(Series1)) and i <= (5 + len(Series1) + len(Series2)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series2[i - 6 - len(Series1)]
                if i > (5 + len(Series1) + len(Series2)) and i <= (5 + len(Series1) + len(Series2) + len(Series3)): tempWorksheet.cell(row = rowIndecies[sheetIndex], column = i).value = Series3[i - 6 - len(Series1) - len(Series2)]

            rowIndecies[sheetIndex] += 1
            core1.save("May23RawData.xlsx")
        counter += 1

        #System Error Recovery
        recoveryFile = open("Recovery.txt","w")
        recoveryFile.write("This are the rowIndecies: " + str(rowIndecies) + "\n")
        recoveryFile.write("This is the current failedIndex: " + str(failedIndex) + "\n")
        recoveryFile.write("Last successful Ticker: " + str(ticker) + "\n")
        recoveryFile.write("Current Exchange: NASDAQ") if selectorBit == 1 else recoveryFile.write("Current Exchange: NYSE")
        recoveryFile.close()
        end = time.time()

        #Impacient programmer pacifier 
        #Change the counter % [Some number] to change number of statements printed
        if (counter % 1 == 0 or counter == 1 or counter == endVal - 1): 
            print("Running ", counter, " of ", endVal - 1, " in NASDAQ ", "| Time Elapsed: ", str(datetime.timedelta(seconds = round(end - start))), \
            " | Cumulative Time Elapsed", str(datetime.timedelta(seconds = round(end - processStartTime)))) \
            if selectorBit == 1 else print("Running ", counter, " of ", endVal - 1, " in NYSE ", "| Time Elapsed: ", str(datetime.timedelta(seconds = round(end - start))), \
            " | Cumulative Time Elapsed", str(datetime.timedelta(seconds = round(end - processStartTime))))
    print("NASDAQ Complete \n") if selectorBit == 1 else print("NYSE Complete \n")
    return rowIndecies, failedIndex

def dataSplitter(input):
    delimiters = ['P/E', 'Forward P/E', 'P/E to S&P500', 'Market CAP', 'Div Yield', '% Held by Insiders', '% Held by Institutions', 
    'Beta', 'PEG', '52w. high/low', 'Avg. Daily Volume', 'Return']
    charOffset = [0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 1]
    inputoffset = [0, 0, 13, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    dotFlag = [False, False, False, False, False, False, False, False, False, False, True, False]
    initialSetRet = list()
    for i in range(len(delimiters)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters[i])], charOffset[i] , dotFlag[i])
        initialSetRet.append(currVal)
        input = input[(indexLoc(input, delimiters[i])+inputoffset[i]):]
    input = input[indexLoc(input, 'linearHighLow'):] # Discard useless data prior to this point
    #ISR contains PE, Forward PE, PE to SP500, MarketCAP(B/T), Dividend, Shares Shorted, Percent Insiders, 
    #Percent Institutions, Beta, Peg Ratio, 52wk h/l, Avg daily shares traded {Additional Definitions at Bottom}
    # numericalDataProcessor(input)
    secondaryRetSet = list()
    highlow = findBackNumber(input[:indexLoc(input, 'Currency:')], 0 , True)
    secondaryRetSet.append(highlow)

    yearsTTM = findBackNumber(input[:indexLoc(input, 'Revenue')], 3 , True)
    secondaryRetSet.append(yearsTTM)

    delimiters2 = ['Earnings per share', 'FCF per share', 'Dividends per share', 'CAPEX per share', 'Book Value per sh.', 'Comm.Shares outs.'
    , 'P/E to S&P500', 'Avg. annual div. yield', 'Revenue (m)', 'Operating margin', 'Depreciation (m)', 'Income tax rate', 'Working capital (m)', 'Equity (m)'
    , 'Return on capital']
    inputoffset2 = [0, 0, 0, 0, 0, 0, 13, 0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(len(delimiters2)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters2[i])], 0 , True)
        secondaryRetSet.append(currVal)
        input = input[(indexLoc(input, delimiters2[i])+inputoffset2[i]):]
    input = input[indexLoc(input, 'Total liabilities'):]# Discard useless data prior to this point
    # print(input)
    #highlow, yrs, Revenue Per Share, Earnings Per Share, FCF Per Share, Dividends, Capex, Book Value/Share, PE Ratio, PE-500, DivYield
    #Revenue, Operating Margin, Net Profit, Net Profit Margin, Equity, ROIC

    finalset = list()
    delimiters3 = ['Total assets', 'Long-term debt', 'Cash and equiv.', 'Goodwill', 'Common stock', 'Working Capital']
    for i in range(len(delimiters3)):
        currVal = findBackNumber(input[:indexLoc(input, delimiters3[i])], 0 , False)
        finalset.append(currVal)
        input = input[indexLoc(input, delimiters3[i]):]
    input = input[indexLoc(input, 'Full-time employees: '):]
    numMinions = findBackNumber(input[:], 0, True)
    finalset.append(numMinions)

    #Total Liabilities, Total Assets, Long-Term Debt, Cash & Equiv., Retained Earnings, Enterprise Value, number of FT workers

    return [initialSetRet, secondaryRetSet, finalset]


def dataCollect(input):
    url = "https://roic.ai/company/" + str(input)
    
    response = requests.get(url)
    contents = BeautifulSoup(response.text, 'html.parser')
    splitter = 'City'
    clipped = contents.get_text().split(splitter)
    return clipped[0]

def genXlSheets(input1, writeloc):
    source = openpyxl.load_workbook(input1)
    e1r = source.active
    writeTo = openpyxl.load_workbook(writeloc)
    target = writeTo.active
    setholder = set()
    #Finding Names of Sections
    for row in e1r.iter_rows(2, e1r.max_row + 1):
        setholder.add(row[9].value)

    for element in setholder:
        # print(writeTo.sheetnames)
        if element in writeTo.sheetnames:
            pass
        elif element == None and 'Unspecified' not in writeTo.sheetnames:
            writeTo.create_sheet('Unspecified')
            writeTo.save(writeloc)
        elif element != None:
            writeTo.create_sheet(element)
            writeTo.save(writeloc)

def indexLoc(input, target):
    location = input.find(target)
    return location

def findBackNumber(input, offset, disableDots):
    returnVal = str()
    dotcounter = 0
    for i in range(len(input) - 1, -1, -1):
        # print('\n XX' + input[i])
        approvedCharacters = {'-', ',', '%', ' ', '/', '(', ')'}
        # if(input[i].isnumeric() or offset > 0 or input[i] in approvedCharacters):
        if(input[i].isnumeric() or offset > 0 or input[i] == '-' or input[i] == ','  or input[i] == '%' or input[i] == ' ' or input[i] == '/' or input[i] == '(' or input[i] == ')'):
            returnVal = input[i] + returnVal
            if (input[i] != '-' and input[i] != ',' and input[i]!= '%' and input[i]!= ' '):
                offset -= 1
        elif (input[i] == '.' and dotcounter == 0):
            returnVal = input[i] + returnVal
            if (disableDots != True):
                dotcounter += 1
        else:
            break
    return returnVal 

if __name__ == "__main__":
    main()


#PEG Ratio: P/E Ratio OVER Earnings Growth Rate
#Beta Number: How stock behaves relative to others. >1 is risky <1 is safe. Unused
