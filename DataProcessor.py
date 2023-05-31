import openpyxl
import time
import datetime
import calendar
import os
import numpy
from math import floor

#Modify Toggles if desired
wipeCurrVerNum, generateSheetToggle = False, True

def main():
    global sheetNames, writeExcelFileName, rawDataBook, core1, rowIndecies, originalStart
    originalStart = time.time()
    print("Starting data processor...")
    writeExcelFileName = generateFileName() #We write to this
    rawDataFileName = findRawDataFileName()
    rawDataBook = openpyxl.load_workbook(rawDataFileName)#Pull data from this 
    rawData = rawDataBook.active
    coreName = "CoreExcelFiles/P2MasterTemplate5.29.23.xlsx"
    core1 = openpyxl.load_workbook(coreName)
    sheetNames = core1.sheetnames
    
    rowIndecies = [3] * (len(sheetNames))
    #Iterate through it, but for now we go for Miscellaneous
    excelWriter()
    end = time.time()
    elapsed = round(end - originalStart)
    print("Total Time Elapsed: ", str(datetime.timedelta(seconds = elapsed)))

def excelWriter():
    global currSheet, yearsTTM, row
    stoppingIndecies = [702, 1220, 1053, 73, 45, 148, 175, 792, 261, 182, 1498, 68, 680]
    counter = 0;
    sheetCounter = 0;
    overallElementCounter = 1;
    for currSheet in sheetNames:
        if currSheet == 'Miscellaneous': #Use for  testing e.g. currSheet == 'Miscellaneous'
            currSheetStartTime = time.time()
            prevSheetEndTime = originalStart
            elementCounter = 1
            #for row in rawDataBook[currSheet].iter_rows(3, stoppingIndecies[sheetCounter] - 1):
            for row in rawDataBook[currSheet].iter_rows(3,44):
                print("Ticker Symbol: ", row[1].value, tickSpaceAdder(row[1].value), "| ", elementCounter, " of ", stoppingIndecies[sheetCounter] - 3,\
                    " in ", currSheet," | ", row[4].value, " | Cumulative Time Elapsed",\
                        str(datetime.timedelta(seconds = round(time.time() - originalStart))), " | ", overallElementCounter, " of ", sum(stoppingIndecies) - 3 * len(stoppingIndecies), " Overall")
                # currSheet = 'Miscellaneous'
                tempWKST = core1[currSheet]
                sheetIndex = sheetNames.index('Miscellaneous')
                #Check 1 and 3 first
                if redFlagsS1(currSheet) and redFlagsS3(currSheet):
                    continueRunning = True
                    #SERIES 2 - More processing than 1 and 3. We don't wanna perform this twice. Errors handled inside here.
                    s2Indecies = [17, 19, 20, 25, 27, 29, 31, 33]
                    s2charSet = [['.'], ['.', '{', '}'], ['.', '(', ')', '{', '}'], ['.', '-', ' '], ['.', '-', ' '], ['.', '-', ' ', '%', '(', ')'], ['.', '-', ' ', '%', '(', ')'], ['.', '-', ' ', '%', '(', ')']]
                    s2Data =  [[0]*18 for i in range(9)]
                    s2ErrorNumStart = 15
                    idNumStart = 1

                    yearsTTM, continueRunning = yearProcessing(18)
                    for i in range (1, len(s2Indecies) + 1):
                        if not continueRunning: break
                        if i == 1: 
                            s2Data[i - 1], s2Data[i], continueRunning = series2Processor(s2Indecies[i - 1], s2ErrorNumStart, idNumStart, s2charSet[i - 1])
                        else:
                            s2Data[i], continueRunning = series2Processor(s2Indecies[i - 1], s2ErrorNumStart, idNumStart, s2charSet[i - 1])
                        idNumStart += 1
                        s2ErrorNumStart += 1
                    #Tag this on for the last one
                    if continueRunning: revenue, continueRunning = series2SmartSplitter(28, 23, s2Data[7], 30)
                    s2Data[8] = revenue
                    if continueRunning: 
                        #TAGS
                        for i in range(1,6):
                            tempWKST.cell(row = rowIndecies[sheetIndex], column = i).value = row[i - 1].value
                        #SERIES 1
                        for i in range(7, 16):
                            x = i - 2
                            if i >= 8: x += 2
                            if i >= 10: x += 1
                            if i >= 15: i += 1
                            if not i == 14:
                                tempWKST.cell(row = rowIndecies[sheetIndex], column = i).value = row[x].value
                            else: 
                                high, low = row[x].value.split('/')
                                tempWKST.cell(row = rowIndecies[sheetIndex], column = i).value = high
                                tempWKST.cell(row = rowIndecies[sheetIndex], column = i + 1).value = low
                        #SERIES 2
                        s3startVal = 197
                        s2rowIndex = 0
                        for i in range(17, s3startVal):
                            if (i - 17) % 18 == 0 and not i == 17: s2rowIndex += 1
                            tempWKST.cell(row = rowIndecies[sheetIndex], column = i).value = s2Data[s2rowIndex][(i - 17) % 18]
                        #SERIES 3
                        counter = 34;
                        for i in range (s3startVal, s3startVal + 5):
                            if i == s3startVal + 3: counter += 2
                            tempWKST.cell(row = rowIndecies[sheetIndex], column = i).value = row[counter].value
                            counter += 1
                        rowIndecies[sheetIndex] += 1
                        if generateSheetToggle: core1.save("ProcessedSheets\\" + monthYR + "\\" + writeExcelFileName)
                elementCounter += 1
                overallElementCounter += 1
            currSheetEndTime = time.time()
            print(currSheet, " Complete in ", str(datetime.timedelta(seconds = round(currSheetEndTime - prevSheetEndTime))), " \n")
            prevSheetEndTime = currSheetStartTime
            sheetCounter += 1

def series2SmartSplitter(revenueIndex, revenueError, netProfitMargin, netProfitIndex):
    if row[revenueIndex].value == None: return None, errorHandler(revenueError, currSheet)
    if row[revenueIndex].value[0] == ')': row[revenueIndex].value = row[revenueIndex].value[1:] # Temporary until infinity issue is fixed
    if row[netProfitIndex].value[0] == ')': row[netProfitIndex].value = row[netProfitIndex].value[1:] # Temporary until infinity issue is fixed
    revOutput = [None] * len(yearsTTM)
    revenueValues = list(); profitValues = list(); possibleValues = list(); outputR = [None] * len(yearsTTM); outputP = [None] * len(yearsTTM)
    rawNumbersRevenue = removeNonNumeric(row[revenueIndex].value, ['(', ')', ',']) 
    rawNumbersProfit = removeNonNumeric(row[netProfitIndex].value, ['(', ')', ',']) 
    if len(rawNumbersRevenue) == 0: return None, errorHandler(revenueError, currSheet)
    if len(rawNumbersProfit) == 0: return None, errorHandler(revenueError, currSheet)

    expNumElemGen = len(netProfitMargin) - netProfitMargin.count(None)
    netProfitMargin = removeNone(netProfitMargin) # This gives us exactly how much we need. 
    #Could add additional check for comma values later.
    try:
        for i in range (0, expNumElemGen):
            # print("\nstartnewrun")
            # print(rawNumbersProfit)
            # print(rawNumbersRevenue)
            possibleValues.clear()
            numberOnlyRevenue = removeNonNumeric(rawNumbersRevenue, [])
            numberOnlyProfit = removeNonNumeric(rawNumbersProfit, [])
            currTargVal = removeNonNumeric(netProfitMargin[i], ['.'])
            # if float(currTargVal) == 0.0: return s2zeroHandler(currRevList=revenueValues)
            #For all -> DO NOT GO OUT OF BOUNDS!
            #Case 1: Revenue has Comma + Net Profit has (). Check 4-6 for R
            if len(rawNumbersRevenue) >= 5 and rawNumbersRevenue[len(rawNumbersRevenue) - 1 - 3] == ',' and rawNumbersProfit[len(rawNumbersProfit) - 1] == ')':
                # print("WE AT 1")
                profitValues.append(rawNumbersProfit[rawNumbersProfit.rindex('('):] ) 
                proposedProfitValue = removeNonNumeric(rawNumbersProfit[rawNumbersProfit.rindex('('):],[])
                rawNumbersProfit = rawNumbersProfit[:rawNumbersProfit.rindex('(')]
                for i in range (4, 7):
                    if i > len(numberOnlyRevenue): break
                    if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:])  == 0: continue
                    # print("i ", i, " ", numberOnlyRevenue [len(str(numberOnlyRevenue)) - i:])
                    possibleValues.append((abs(int( proposedProfitValue)/int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i))
                numDigits = min(possibleValues)[1]
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numDigits:])#-1 removed on 5.30.23 at 1:02 AM ET
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numDigits - 1]
            #Case 2: Revenue has no comma + Net Profit has (). Check 1-3 for R
            elif rawNumbersProfit[len(rawNumbersProfit) - 1] == ')':
                # print("WE AT 2")
                profitValues.append(rawNumbersProfit[rawNumbersProfit.rindex('('):] )
                proposedProfitValue = removeNonNumeric(rawNumbersProfit[rawNumbersProfit.rindex('('):],[])
                rawNumbersProfit = rawNumbersProfit[:rawNumbersProfit.rindex('(')]
                # print(proposedProfitValue)
                for i in range (1, 4):
                    if i > len(numberOnlyRevenue): break;
                    if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) == 0: continue
                    # print("i ", i, " ", numberOnlyRevenue [len(str(numberOnlyRevenue)) - i:])
                    possibleValues.append((abs(int( proposedProfitValue)/int(numberOnlyRevenue \
                        [len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i))
                numDigits = min(possibleValues)[1]
                # print(possibleValues, numDigits)
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numDigits:])
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numDigits]
            #Case 3: Revenue has Comma + Net Profit has comma. Check 4-6 for both
            elif len(rawNumbersRevenue) >= 5 and rawNumbersRevenue[len(rawNumbersRevenue) - 1 - 3] == ',' \
                and  len(rawNumbersProfit) >= 5 and rawNumbersProfit[len(rawNumbersProfit) - 1 - 3] == ',':
                # print("WE AT 3")
                for j in range (4, 7): # Profit
                    for i in range (4, 7): # Revenue
                        if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) == 0: continue
                        if i <= len(numberOnlyRevenue) and j <= len(numberOnlyRevenue):
                            possibleValues.append((abs(int( numberOnlyProfit[len(str(numberOnlyProfit)) - j:])/ \
                                int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i, j))
                dummy, numRevenue, numProfit = min(possibleValues)
                # print(numRevenue, numProfit)
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numRevenue:])#-1 removed on 5.30.23 at 1:02 AM ET
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numRevenue - 1]
                profitValues.append(numberOnlyProfit[len(numberOnlyProfit) - numProfit:])#-1 removed on 5.30.23 at 1:02 AM ET
                rawNumbersProfit = rawNumbersProfit[:len(rawNumbersProfit) - numProfit - 1]
            #Case 4: Revenue has Comma + Net Profit has no comma. Check 4-6 for R and 1-3 for NP
            elif len(rawNumbersRevenue) >= 5 and rawNumbersRevenue[len(rawNumbersRevenue) - 1 - 3] == ',':
                # print("WE AT 4")
                for j in range (1, 4): # Profit
                    for i in range (4, 7): # Revenue
                        if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) == 0: continue
                        if i <= len(numberOnlyRevenue) and j <= len(numberOnlyRevenue):
                            # print("i ", i, " j ", j, " ", int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]))
                            # print("my curr targ ", currTargVal)
                            possibleValues.append((abs(int( numberOnlyProfit[len(str(numberOnlyProfit)) - j:])/ \
                                int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i, j))
                dummy, numRevenue, numProfit = min(possibleValues)
                # print(possibleValues, numRevenue, numProfit)
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numRevenue:])#Add in -1 if you're not doing number only
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numRevenue - 1]
                profitValues.append(numberOnlyProfit[len(numberOnlyProfit) - numProfit:])
                rawNumbersProfit = rawNumbersProfit[:len(rawNumbersProfit) - numProfit]
            #Case 5: Revenue has no comma + Net Profit has comma. Check 1-3 for R and 4-6 for NP
            elif len(rawNumbersProfit) >= 5 and rawNumbersProfit[len(rawNumbersProfit) - 1 - 3] == ',':
                # print("WE AT 5")
                for j in range (4, 7): # Profit
                    for i in range (1, 4): # Revenue
                        if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) == 0: continue
                        if i <= len(numberOnlyRevenue) and j <= len(numberOnlyRevenue):
                            # print("i ", i, " j ", j, " ", numberOnlyProfit[len(str(numberOnlyProfit)) - j:])
                            possibleValues.append((abs(int( numberOnlyProfit[len(str(numberOnlyProfit)) - j:])/ \
                                int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i, j))
                dummy, numRevenue, numProfit = min(possibleValues)
                # print(numRevenue, numProfit)
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numRevenue:])
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numRevenue]
                profitValues.append(numberOnlyProfit[len(numberOnlyProfit) - numProfit:])#-1 removed on 5.30.23 at 1:02 AM ET
                rawNumbersProfit = rawNumbersProfit[:len(rawNumbersProfit) - numProfit - 1]
            #Case 6: Revenue has no comma + Net Profit has no comma. Check 1-3 for both
            else:
                # print("WE AT 6")
                for j in range (1, 4): # Profit
                    for i in range (1, 4): # Revenue
                        if int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) == 0: continue
                        if i <= len(numberOnlyRevenue) and j <= len(numberOnlyRevenue):
                            # print("i ", i, " j ", j, " ", numberOnlyProfit[len(str(numberOnlyProfit)) - j:])
                            possibleValues.append((abs(int( numberOnlyProfit[len(str(numberOnlyProfit)) - j:])/ \
                                int(numberOnlyRevenue[len(str(numberOnlyRevenue)) - i:]) - float(currTargVal)/100), i, j))
                dummy, numRevenue, numProfit = min(possibleValues)
                # print(possibleValues, numRevenue, numProfit)
                revenueValues.append(numberOnlyRevenue[len(numberOnlyRevenue) - numRevenue:])
                rawNumbersRevenue = rawNumbersRevenue[:len(rawNumbersRevenue) - numRevenue]
                profitValues.append(numberOnlyProfit[len(numberOnlyProfit) - numProfit:])
                rawNumbersProfit = rawNumbersProfit[:len(rawNumbersProfit) - numProfit]
    except:
        return s2zeroHandler(currRevList=revenueValues)
    
    print(rawNumbersRevenue)
    if len(rawNumbersRevenue) > 0: return s2zeroHandler(currRevList=revenueValues)

    for i in range(0, len(yearsTTM)):
        if yearsTTM[i] == 1 and len(revenueValues) > 0:
            outputR[i] = revenueValues.pop(0)
            outputP[i] = profitValues.pop(0)
    # print(outputP)
    # print(outputR)
    #return outputR, outputP, True
    return outputR, True

def s2zeroHandler(currRevList):
    for i in range (0, len(yearsTTM)):
        if i < len(currRevList):
            currRevList[i] = str(currRevList[i]) + 'X'
        else:
            currRevList.append('X')
    return currRevList, True

def removeNone(currlist):
    temp = list()
    for i in range(0, len(currlist)):
        if not currlist[i] == None:
            temp.append(currlist[i])
    return temp

def series2Processor(rowIndex, errorNum, idNum, additionalSet):
    if row[rowIndex].value == None: return None, errorHandler(errorNum, currSheet)
    if row[rowIndex].value[0] == ')' and idNum == 3: row[rowIndex].value = row[rowIndex].value[1:] # Temporary until infinity issue is fixed
    if row[rowIndex].value[0:2] == ')%' and idNum == 8: row[rowIndex].value = row[rowIndex].value[2:] # Temporary until infinity issue is fixed

    output = [None] * len(yearsTTM)
    if idNum == 1: output2 = [None] * len(yearsTTM)
    individualValues = list()
    rawNumbers = removeNonNumeric(row[rowIndex].value, additionalSet) 
    if len(rawNumbers) == 0 and not idNum == 5: return None, errorHandler(errorNum, currSheet)
    while len(rawNumbers) > 0:
        individualValues, rawNumbers = series2ProcessorCondHelper(idNum, individualValues, rawNumbers)
    for i in range(0, len(yearsTTM)):
        if yearsTTM[i] == 1 and len(individualValues) > 0:
            output[i] = individualValues.pop(len(individualValues) - 1)
            if idNum == 1: output2[i] = individualValues.pop(len(individualValues) - 1)
    if idNum == 1: return output, output2, True
    return output, True

def series2ProcessorCondHelper(idNum, individualValues, rawNumbers):
    if idNum in [6, 7, 8]:
        if rawNumbers[0] == '-':
            individualValues.append(str(0.0))
            rawNumbers = rawNumbers[3:]
        else:
            individualValues.append(rawNumbers[0:rawNumbers.index('%') + 1])
            rawNumbers = rawNumbers[rawNumbers.index('%') + 1:]
    if idNum in [4]:
        if rawNumbers[0] == '(':
            individualValues.append(rawNumbers[0:rawNumbers.index(')') + 1])
            rawNumbers = rawNumbers[rawNumbers.index(')') + 1:]
    if idNum in [4, 5]:
        if rawNumbers[0] == '-':
            individualValues.append(str(0.0))
            rawNumbers = rawNumbers[3:]
        else:
            individualValues.append(rawNumbers[0:rawNumbers.index('.') + 2])
            rawNumbers = rawNumbers[rawNumbers.index('.') + 2:]
    if idNum in [3]:
        if rawNumbers[0] == '(':
            individualValues.append(rawNumbers[0:rawNumbers.index(')') + 1])
            rawNumbers = rawNumbers[rawNumbers.index(')') + 1:]
        elif rawNumbers[0] == '{':
            individualValues.append(rawNumbers[0:rawNumbers.index('}') + 1])
            rawNumbers = rawNumbers[rawNumbers.index('}') + 1:]
        else:
            individualValues.append(rawNumbers[0:rawNumbers.index('.') + 3])
            rawNumbers = rawNumbers[rawNumbers.index('.') + 3:]
    if idNum in [2]:
        if rawNumbers[0] == '{':
            individualValues.append(rawNumbers[0:rawNumbers.index('}') + 1])
            rawNumbers = rawNumbers[rawNumbers.index('}') + 1:]
    if idNum in [1, 2]:
        individualValues.append(rawNumbers[0:rawNumbers.index('.') + 3])
        rawNumbers = rawNumbers[rawNumbers.index('.') + 3:]
    return individualValues, rawNumbers

def yearProcessing(rowIndex):
    errorNum = 14
    markers = [0] * 18
    startVal = 2023
    for i in range(0, 18):
        if i == 0: markers[i] = 'TTM' in row[rowIndex].value
        else: 
            markers[i] = str(startVal) in row[rowIndex].value
            startVal -= 1
    if markers.count(1) == 0: 
        return markers, errorHandler(errorNum, currSheet)
    return markers, True

def redFlagsS1(sheetName):
    #41
    #PE, MKTCAP, Share Short, %Insider, %Institution, 52HighLow, DailyTrade - Series 1
    performanceValues = [5, 8, 10, 11, 12, 15, 16] 
    additionalSet = ['-', '.']
    performanceLength = list()
    errorNum = 0
    for i in performanceValues:
        if row[i].value == None:
            return errorHandler(782, sheetName)
        performanceLength.append(len(removeNonNumeric(row[i].value, additionalSet)))
        if i == 5 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there PE
            return errorHandler(errorNum, sheetName)
        if i == 5 and float(removeNonNumeric(row[i].value, additionalSet)) < -100: #PE Over 300
            return errorHandler(errorNum + 1, sheetName)
        if i == 5 and float(removeNonNumeric(row[i].value, additionalSet)) > 300: #PE Under -100
            return errorHandler(errorNum + 2, sheetName)
        if i == 8 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there MKTCAP
            return errorHandler(errorNum + 3, sheetName)
        if i == 10 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there SHARE SHORT
            return errorHandler(errorNum + 4, sheetName)
        if i == 10 and float(removeNonNumeric(row[i].value, additionalSet)) > 20.0: #Short percentage is over 20 percent...
            return errorHandler(errorNum + 5, sheetName)
        if i == 11 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there Percent Insider
            return errorHandler(errorNum + 6, sheetName)
        if i == 12 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there Institution %
            return errorHandler(errorNum + 7, sheetName)
        if i == 12 and float(removeNonNumeric(row[i].value, additionalSet)) > 95: #Institution Value too %
            return errorHandler(102, sheetName)
        if i == 15 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there high low
            return errorHandler(errorNum + 8, sheetName)
        if i == 16 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there daily trade volume
            return errorHandler(errorNum + 9, sheetName)
        if i == 16 and float(removeNonNumeric(row[i].value, additionalSet)) < .15: #Dailhy trade volume below 150,000
            return errorHandler(103, sheetName)
        if i == 16 and performanceLength.count(0) >= 6: # 6 or more fields missing in Series
            return errorHandler(errorNum + 10, sheetName)
    return True

#assets less than 1000 eliminated maybe change this
def redFlagsS3(sheetName):
    performanceValues = [34, 35, 40] 
    additionalSet = ['-', '.'] 
    #Total Liabilities, Total Assets and Number of Employees
    errorNum = 11
    errorNum2 = 100
    for i in performanceValues: 
        if row[i].value == None:
            return errorHandler(782, sheetName)
        if i == 35 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there Total Liabilities
            return errorHandler(errorNum, sheetName)
        if i == 36 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Nothing is there Total Assets
            return errorHandler(errorNum + 1, sheetName)
        if i == 40 and len(removeNonNumeric(row[i].value, additionalSet)) == 0: #Number of employees missing
            return errorHandler(errorNum2, sheetName) 
        if i == 40 and int(removeNonNumeric(row[i].value, additionalSet)) < 50: #Short percentage is over 20 percent...
            return errorHandler(errorNum2 + 1, sheetName)
        # if i == 36 and float(removeNonNumeric(row[i].value, additionalSet)) < 10: #Assets less than $10 Million
        #     return errorHandler(errorNum + 2, sheetName)
    return True

def removeNonNumeric(input, additionalSet):
    output = ""
    approvedSet = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    counter = 0
    for i in range(0, len(input)):
        if (input[i] in approvedSet):
            counter += 1
            output+= input[i]
        if (input[i] in additionalSet):
            output+= input[i]
    if counter == 0: return ""
    return output

def errorHandler(flag, industry):
    tempWKST = core1['ELIMINATED']
    sheetIndex = sheetNames.index('ELIMINATED')
    flagComment = getErrorCode(flag)
    for i in range(1,7):
        x = i
        if i == 5: 
            tempWKST.cell(row = rowIndecies[sheetIndex] - 1, column = i).value = industry
        else:
            if i > 5: x -= 1
            tempWKST.cell(row = rowIndecies[sheetIndex] - 1, column = i).value = row[x - 1].value
    tempWKST.cell(row = rowIndecies[sheetIndex] - 1, column = 7).value = flagComment
    rowIndecies[sheetIndex] += 1
    if generateSheetToggle: core1.save("ProcessedSheets\\" + monthYR + "\\" + writeExcelFileName)
    return False;

def findRawDataFileName():
    global monthYR
    recoveryFile = open('Recovery.txt')
    recoveryLines = recoveryFile.readlines()
    if (len(recoveryLines) < 3):
        print("ERROR: Please check that your Recovery.txt file has been generated properly.")
        monthYR = "ERROR"
        versionNum = 0
    else:
        monthYR = recoveryLines[4].split(" ")[1].strip()
        versionNum = recoveryLines[5].split(" ")[1].strip()
    return str(monthYR) + "RawDataV" + str(versionNum) + ".xlsx"

def generateFileName():
    recoveryFile = open('Recovery.txt')
    recoveryLines = recoveryFile.readlines()
    recoveryFile.close()
    MonthYear = recoveryLines[4].split(" ")[1].strip()
    filePath = "C://Users//User//Documents//GitHub//Automated-Stock-Selector//ProcessedSheets"
    directoryFiles = os.listdir(filePath)
    if not (MonthYear in directoryFiles): #Create the folder if it isn't already there.
        folderPath = os.path.join(filePath, MonthYear)
        os.mkdir(folderPath)
    filePath = os.path.join(filePath, MonthYear)
    versionNumber = -1
    #Initial Proposed file.
    if len(recoveryLines) == 7: 
        recoveryFile = open("Recovery.txt","a")
        recoveryFile.write("\nDataProcessVersion: " + str(1))
        versionNumber = 1
        recoveryFile.close()
    else: 
        writeToRecovery(0, None, wipeCurrVerNum)
        versionNumber = (int(recoveryLines[7].split(" ")[1].strip()) + 1) if not wipeCurrVerNum else 1
    proposedFileName = "ProcessedDataV" + str(versionNumber) + ".xlsx"

    directoryFiles = os.listdir(filePath)
    #Check this proposed fileName
    contentModified = False
    while proposedFileName in directoryFiles:
        contentModified = True
        versionNumber += 1
        proposedFileName = "ProcessedDataV" + str(versionNumber) + ".xlsx"
    if contentModified:
        writeToRecovery(1, versionNumber, False)
    return proposedFileName

def writeToRecovery(toggle, versionNumber, wipeCurrVerNum):
    recoveryFile = open('Recovery.txt', 'r')
    recoveryLines = recoveryFile.readlines()
    if not wipeCurrVerNum:
        recoveryLines[7] = "DataProcessVersion: " + \
            (str((int(recoveryLines[7].split(" ")[1].strip()) + 1)) if toggle == 0 else str(versionNumber))
    else:
        recoveryLines[7] = "DataProcessVersion: " + str(1)
    recoveryFile = open('Recovery.txt', 'w')
    for i in range(0, len(recoveryLines)):
        recoveryFile.write(str(recoveryLines[i]))
    recoveryFile.close()

def getErrorCode(input):
    switch = {
        0: "E0: P/E Ratio is missing",
        1: "E1: P/E Ratio is below the cutoff of -100",
        2: "E2: P/E Ratio is above the cutoff of +300",
        3: "E3: Missing Market Cap Value",
        4: "E4: Missing Shares Shorted Value",
        5: "E5: Shares Shorted Value over 20 percent",
        6: "E6: Missing Percent of Insiders",
        7: "E7: Percent held by institutions is missing",
        8: "E8: 52 Week High and Low missing",
        9: "E9: Daily trade volume missing",
        10: "E10: Missing 6 or more of Series 1 fields",
        11: "E11: Missing Total Liabilities",
        12: "E12: Missing Total Assets",
        13: "E13: Total assets are below $1000",
        14: "E14: Years/TTM Missing",
        15: "E15: Yearly High/Low Values Missing",
        16: "E16: RPS Values Missing",
        17: "E17: EPS Values Missing",
        18: "E18: P/E Values Missing",
        19: "E19: Dividend Yield Value Missing",
        34: "E34: Missing 6 or more of Series 2 fields",
        35: "E35: Series 3 Value is missing",
        36: "E36: ROIC value is below (last 3, last 5 avg and all avg)",
        100: "E100: Number of Employees Missing",
        101: "E101: Number of Employees below 50",
        102: "E102: % Held by institutions over 100",
        103: "E103: Daily trading volume below 150k",
        782: "E782: Stock has Element set to 'None'",
        }

def tickSpaceAdder(tick):
    tickOffset = 6 - len(str(tick))
    output = ""
    output += ' ' * tickOffset
    return output

    return switch.get(input, "")
if __name__ == "__main__":
    main()